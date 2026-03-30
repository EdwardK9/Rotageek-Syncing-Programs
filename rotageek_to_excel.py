import requests
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

# --- CONFIG ---
URL = "https://screwfix.rotageek.com/api/graphql-userschedules"
PINK_FILL = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")


class ShiftSelector(tk.Toplevel):
    def __init__(self, parent, items):
        super().__init__(parent)
        self.title("Select Shifts to Sync")
        self.geometry("450x650")
        self.result = []

        # Handle the "X" button click
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        # Header and Bulk Actions
        header_frame = tk.Frame(self, pady=10)
        header_frame.pack(fill="x")
        tk.Label(header_frame, text="Verify shifts to sync:", font=("Arial", 10, "bold")).pack()

        bulk_frame = tk.Frame(self)
        bulk_frame.pack(fill="x", pady=5)
        tk.Button(bulk_frame, text="Select All", command=self.select_all).pack(side="left", padx=20)
        tk.Button(bulk_frame, text="Deselect All", command=self.deselect_all).pack(side="left")

        # Scrollable area
        container = ttk.Frame(self)
        container.pack(expand=True, fill="both", padx=5, pady=5)

        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        self.vars = []
        for item in items:
            s_dt = datetime.fromisoformat(item['start'].replace('Z', ''))
            e_dt = datetime.fromisoformat(item['end'].replace('Z', ''))
            is_leave = item.get('isLeave') or 'leaveType' in item

            type_label = "🌸 ANNUAL LEAVE" if is_leave else f"🕒 {s_dt.strftime('%H:%M')} - {e_dt.strftime('%H:%M')}"
            display_text = f"{s_dt.strftime('%a %d %b')}: {type_label}"

            # Using self as master to ensure the state is preserved
            var = tk.BooleanVar(self, value=True)

            cb = tk.Checkbutton(self.scrollable_frame, text=display_text, variable=var, anchor="w",
                                font=("Consolas", 10))
            cb.pack(fill="x", padx=10, pady=2)
            self.vars.append((var, item))

        canvas.pack(side="left", expand=True, fill="both")
        scrollbar.pack(side="right", fill="y")

        # Footer Buttons
        btn_frame = tk.Frame(self, pady=15)
        btn_frame.pack(fill="x")
        tk.Button(btn_frame, text="Sync Selected to Excel", command=self.confirm, bg="#28a745", fg="white", width=20,
                  height=2).pack(side="right", padx=20)
        tk.Button(btn_frame, text="Cancel", command=self.on_close).pack(side="right")

    def select_all(self):
        for var, _ in self.vars: var.set(True)

    def deselect_all(self):
        for var, _ in self.vars: var.set(False)

    def on_close(self):
        """Kills the entire script if the window is closed."""
        self.destroy()
        sys.exit()

    def confirm(self):
        self.result = [item for var, item in self.vars if var.get()]
        self.destroy()


def get_credentials():
    root = tk.Tk();
    root.withdraw()
    instructions = (
        "HOW TO GET CREDENTIALS:\n"
        "1. Login to Rotageek (Chrome/Edge).\n"
        "2. Press F12 -> Network tab.\n"
        "3. Refresh the page.\n"
        "4. Click on 'graphql-userschedules' in the list.\n"
        "5. In 'Request Headers', find 'cookie' and 'requestverificationtoken'."
    )
    cookie = simpledialog.askstring("Step 1: Cookie", f"{instructions}\n\nPaste 'cookie':")
    if not cookie: sys.exit()
    token = simpledialog.askstring("Step 2: Token", "Paste 'requestverificationtoken':")
    if not token: sys.exit()
    return cookie, token


def update_excel():
    cookie, token = get_credentials()

    root = tk.Tk();
    root.withdraw()
    path = filedialog.askopenfilename(title="Select Screwfix Excel", filetypes=[("Excel", "*.xlsx")])
    if not path: sys.exit()

    now = datetime.now()
    start_q = (now - timedelta(days=7)).strftime('%Y-%m-%dT00:00:00')
    end_q = (now + timedelta(days=28)).strftime('%Y-%m-%dT23:59:59')

    headers = {"content-type": "application/json", "cookie": cookie, "requestverificationtoken": token}
    payload = {
        "query": """query { 
            dateRangeSchedule(start: \"""" + start_q + """\", end: \"""" + end_q + """\") { 
                journals { start end untimedUnpaidBreakDuration isLeave }
                absences { start end leaveType { name } }
            } 
        }"""
    }

    try:
        r = requests.post(URL, headers=headers, json=payload)
        data = r.json().get('data', {}).get('dateRangeSchedule', {})
        raw_items = data.get('journals', []) + data.get('absences', [])
        raw_items.sort(key=lambda x: x['start'])
    except:
        messagebox.showerror("Error", "Check your Token/Cookie.")
        sys.exit()

    # Show the Selector Window
    selector = ShiftSelector(root, raw_items)
    root.wait_window(selector)
    selected_items = selector.result

    if not selected_items:
        sys.exit()

    try:
        wb = load_workbook(path, data_only=False)
    except PermissionError:
        messagebox.showerror("File Error", "Please CLOSE the Excel file.")
        sys.exit()

    updated_count = 0
    for item in selected_items:
        s_dt = datetime.fromisoformat(item['start'].replace('Z', ''))
        e_dt = datetime.fromisoformat(item['end'].replace('Z', ''))
        shift_date = s_dt.date()
        target_sheet = next((n for n in wb.sheetnames if s_dt.strftime('%b') in n and s_dt.strftime('%Y') in n), None)
        if not target_sheet: continue
        ws = wb[target_sheet]

        for row_idx in range(1, 50):
            val = ws.cell(row=row_idx, column=2).value
            if (isinstance(val, datetime) and val.date() == shift_date) or (val and s_dt.strftime('%d') in str(val)):

                for m_range in list(ws.merged_cells.ranges):
                    if m_range.min_row <= row_idx <= m_range.max_row and m_range.min_col <= 4:
                        ws.unmerge_cells(range_string=m_range.coord)

                if item.get('isLeave') or 'leaveType' in item:
                    ws.merge_cells(start_row=row_idx, start_column=3, end_row=row_idx, end_column=4)
                    cell = ws.cell(row=row_idx, column=3)
                    cell.value, cell.fill = "Annual Leave", PINK_FILL
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(row=row_idx, column=5).value = (e_dt - s_dt).total_seconds() / 3600
                else:
                    ws.cell(row=row_idx, column=3).value = s_dt.strftime('%H:%M')
                    ws.cell(row=row_idx, column=4).value = e_dt.strftime('%H:%M')
                    ws.cell(row=row_idx, column=6).value = item.get('untimedUnpaidBreakDuration', 0)

                updated_count += 1
                break

    wb.save(path)
    messagebox.showinfo("Success", f"Sync Complete! Updated {updated_count} items.")


if __name__ == "__main__":
    update_excel()